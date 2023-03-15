from tkinter import filedialog as fd
from tkinter import ttk
from tkinter import messagebox
import tkinter as tk
import numpy as np
import os
import os.path
import time
import shutil
import datetime
from datetime import datetime, date, timedelta 
import docx
import pandas as pd
import openpyxl as ox
from openpyxl.styles import (
                        PatternFill, Border, Side, 
                        Alignment, Font, GradientFill
                        )

def set_folder_xlsx():
        global file_names_xlsx_zip
        file_names_xlsx_zip = fd.askopenfilename(multiple=True, initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ\разбор в МО\В МО')


def create_send_list():
    df_base = pd.read_excel(file_xlsx, sheet_name='Иды', dtype=str)
    df_base_new = df_base.loc[df_base['Отработано'].isna()]
    df_base_new = df_base_new.loc[:,(['Внесено', 'Комментарии', 'Фамилия пациента', 'Имя пациента', 'Отчество', 'Дата рождения',
 'Полис ОМС', 'СНИЛС пациента из документа', 'Наменование МО', 'Дата вакцинации оцифровка', 'Кратность вакцинации', 
 'Препарат вакцины', 'Серия и контрольный номер'])]
    df_base_new.rename(columns = {'Фамилия пациента':'Фамилия', 'Дата вакцинации оцифровка':'Дата вакцинации',
                                   'Кратность вакцинации':'Кратность', 'Препарат вакцины':'Препарат',
                                   'Серия и контрольный номер':'Серия'}, inplace = True )
    
    df_base_new['Препарат'] = df_base_new['Препарат'].str.split(')').str[0] + ')'

    df_base_new.insert(0, 'Код', '')
    df_base_new['Код'] = np.arange(len(df_base_new))+1
    df_base_new['Запрос'] = ''
    df_base_new['способ введения вакцины: внутримышечно или интраназально'] = ''
    df_base_new['Кратность\n(V1 или V2)'] = ''
    df_base_new['Компонент \n(I компонент или II компонент)'] = ''
    df_base_new['Отработка\nВписать сюда что требуется в графе "Запрос"'] = ''
    df_base_new['ФИО сотрудника'] = ''
    df_base_new['Должность сотрудника'] = ''
    df_base_new['МО сотрудника'] = ''
    
    # Подстановка коротких названий МО в выгрузку
    df_name_mo = pd.read_excel(file_xlsx, sheet_name='экспертное', dtype=str)
    df_unic_mo = df_name_mo.drop_duplicates (subset=['ИСХ МО']).dropna(subset=['ИСХ МО'])
    df_mo = pd.Series(df_unic_mo['МО'].to_list(), df_unic_mo['ИСХ МО'])
    df_base_new['МО'] = df_base_new['Наменование МО']
    df_base_new['МО'] = df_base_new['МО'].map(df_mo)
    
    df_base_new['Написать способ введения, кратность и компонент вакцины /n'] = ''

    wb = ox.load_workbook(filename=file_xlsx, read_only=False)
    today = datetime.now().date().strftime("%d.%m.%y")
    ws = wb.create_sheet(f'отработка {today}', 3)

    thins = Side(border_style="thin", color="000000")

    for i, value in enumerate(list(df_base_new), 1):
        ws.cell(row=1, column=i).value = value
        ws.cell(row=1, column=i).font = Font(bold=True)
        ws.cell(row=1, column=i).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        ws.cell(row=1, column=i).border = Border(top=thins, bottom=thins, left=thins, right=thins)

    for ir in range(0, len(df_base_new)):
        for ic in range(0, len(df_base_new.iloc[ir])):
            ws.cell(row=2 + ir,column=1 + ic).value = df_base_new.iloc[ir][ic]
            ws.cell(row=2 + ir,column=1 + ic).border = Border(top=thins, bottom=thins, left=thins, right=thins)
            if 1 + ic != 3:
                ws.cell(row=2 + ir,column=1 + ic).alignment = Alignment(vertical='center', wrapText=True)
            else:
                ws.cell(row=2 + ir,column=1 + ic).alignment = Alignment(vertical='center')      
         
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 4
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 13
    ws.column_dimensions['I'].width = 13
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 4
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['O'].width = 50
    ws.column_dimensions['S'].width = 40
      
    wb.save(file_xlsx)

    ws = wb['Иды']
    for ir in range(2, ws.max_row + 1):
        if ws.cell(row=ir, column=1).value is None:
            ws.cell(row=ir, column=1).value = 'отдано'
            ws.cell(row=ir, column=1).fill = PatternFill('solid', fgColor='E2EFDA')
        
    wb.save(file_xlsx)
    wb.close()
    os.startfile(file_xlsx)


def Copy_file():
    todays_date = date.today()
    todays_date = todays_date.strftime("%d.%m.%Y")
    addname = ' ' + todays_date

    global folder_xlsx

    if folder_xlsx == '':
            folder_xlsx = os.path.dirname(file_names_xlsx_zip[0])
    for file_names in os.listdir(folder_xlsx):
        base_name = os.path.basename(file_names)
        base_name = base_name[:-5]
        shutil.copyfile(os.path.join(folder_docx, base_name + '.docx'), os.path.join(folder_xlsx, base_name + addname +'.docx'))   
        newfilename = os.path.splitext(folder_xlsx + file_names)[0] + addname + os.path.splitext(file_names)[1]
        os.rename(folder_xlsx + file_names, newfilename)

    def replace_text():
        old_text = find_entry.get()
        new_text = replace_entry.get()
        directory = folder_xlsx
        for filename in os.listdir(directory):
            if filename.endswith(".docx"):
                doc = docx.Document(os.path.join(directory, filename))
                for paragraph in doc.paragraphs:
                    for run in paragraph.runs:
                        if old_text in run.text:
                            font = run.font
                            run.text = run.text.replace(old_text, new_text)
                            run.font.size = font.size
                doc.save(os.path.join(directory, filename))

        messagebox.showinfo("Info", "Работа сделана!!")
        os.startfile(folder_xlsx)

        root.destroy()
        
    innerroot = tk.Toplevel(root)
    root.withdraw()
    innerroot.title('Replace text in .docx files')
    
    find_label = tk.Label(innerroot, text="Find:")
    find_entry = tk.Entry(innerroot)
    find_entry.insert(0, 'Дата_ответа')

    replace_label = tk.Label(innerroot, text="Replace:")
    replace_entry = tk.Entry(innerroot)
    replace_entry.insert(0, tomorrow)

    replace_button = tk.Button(innerroot, text="Replace", command=replace_text)

    find_label.pack()
    find_entry.pack()
    replace_label.pack()
    replace_entry.pack()
    replace_button.pack()

    innerroot.mainloop()
    
def sep_base_mo():
    #Проверка и создание папки
    today = pd.Timestamp(datetime.now()).date().strftime("%d.%m.%y")
    global folder_xlsx
    folder_xlsx = os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today + '\\отправка\\'
    if not os.path.exists(folder_xlsx):
        if not os.path.exists(os.path.dirname(file_xlsx) + '\\разбор в МО\\'):
            os.mkdir(os.path.dirname(file_xlsx) + '\\разбор в МО\\')
        else:
            if not os.path.exists(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\'):
                os.mkdir(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\')
            else:
                if not os.path.exists(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today):
                    os.mkdir(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today)
                else:
                    if not os.path.exists(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today + '\\отправка\\'):
                        os.mkdir(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today + '\\отправка\\')
                    if not os.path.exists(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today + '\\прием\\'):
                        os.mkdir(os.path.dirname(file_xlsx) + '\\разбор в МО\\В МО\\' + today + '\\прием\\')

        
    # чтение исходного файла
    df_base_mo = pd.read_excel(file_xlsx, sheet_name='отработка ' + today, dtype=str)
    df_send = df_base_mo.loc[:,('Код', 'Фамилия', 'Имя пациента', 'Отчество', 'Дата рождения', 'Полис ОМС', 'Наменование МО',
 'Дата вакцинации', 'Кратность', 'Препарат', 'Серия', 'Запрос', 'способ введения вакцины: внутримышечно или интраназально',
 'Кратность\n(V1 или V2)', 'Компонент \n(I компонент или II компонент)', 'Отработка\nВписать сюда что требуется в графе "Запрос"',
 'ФИО сотрудника', 'Должность сотрудника', 'МО сотрудника', 'МО')]

    # разделение датафрейма по столбцу "МО"
    grouped = df_send.groupby('МО')

    # сохранение каждой группы как отдельный файл
    for name, group in grouped:
        group.iloc[:,:-1].to_excel(f'{folder_xlsx}{name}.xlsx', index=False)

    # Редактирование созданных файлов
    name_mo = df_send['МО'].drop_duplicates()
    for name in name_mo:
        file_MO = f'{folder_xlsx}{name}.xlsx'
        wb = ox.load_workbook(filename=file_MO, read_only=False)
        ws = wb.active
        thins = Side(border_style="thin", color="000000")  
        for i in ('ABCDEFGHIJKLMNOPQRS'):
            for rw in range(2, ws.max_row + 1):
                ws[i + str(rw)].alignment = Alignment(vertical='center', wrapText=True)
                ws[i + str(rw)].border = Border(top=thins, bottom=thins, left=thins, right=thins)
            ws[i + '1'].font = Font(bold=True)
            ws[i + '1'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
         
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['G'].width = 16
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 4
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['L'].width = 50
        ws.column_dimensions['P'].width = 40
    
        wb.save(file_MO)
        wb.close()
    Copy_file()

def join_xlsx_mo():
    folder_report = fd.askdirectory(title='Выберите папку с ответами МО', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ\разбор в МО\В МО')
    df_report_tot = pd.DataFrame()
    #Объединяем ответы в одну таблицу
    for filename in os.listdir(folder_report):
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            df_report = pd.read_excel(os.path.join(folder_report, filename), dtype=str)
            df_report_tot = df_report_tot.append(df_report, ignore_index=True, sort=False)

    df_report_tot_short = df_report_tot.loc[~df_report_tot['Фамилия'].isna() & ~df_report_tot['Имя пациента'].isna() & ~df_report_tot['Полис ОМС'].isna()]
    df_report_tot_short['Код'] = df_report_tot_short['Код'].astype(int)

    df_xlsx = pd.read_excel(file_xlsx, sheet_name=3, dtype=str)
    df_merged = df_xlsx.loc[:,('Фамилия', 'Имя пациента', 'Отчество', 'Дата рождения',
                               'Полис ОМС', 'Наменование МО',
                               'Дата вакцинации', 'Кратность', 'Препарат', 'Серия'
                               )].merge(df_report_tot_short, 
                                        how='left',
                                        validate='1:1'
                                        )
    df_to_file_xlsx = df_merged.iloc[:, 12:]
    wb = ox.load_workbook(filename=file_xlsx, read_only=False)
    ws = wb.worksheets[3]
    thins = Side(border_style="thin", color="000000")
    #Заполнение данными из df
    for ir in range(0, len(df_to_file_xlsx)):
        for ic in range(0, len(df_to_file_xlsx.iloc[ir])):
            ws.cell(row=2 + ir,column=16 + ic).value = df_to_file_xlsx.iloc[ir][ic]
            ws.cell(row=2 + ir,column=16 + ic).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws.cell(row=2 + ir,column=16 + ic).border = Border(top=thins, bottom=thins, left=thins, right=thins)
                
    wb.save(file_xlsx)
    wb.close()   

    root.destroy()
    os.startfile(file_xlsx)

            
#Проверка наличие шаблонов сопроводиловок
folder_docx = "C:\\Users\\User\\Емиас\\Корректировка ЕГИСЗ\\разбор в МО\\В МО\\шаблоны\\"
if not os.path.exists(folder_docx):
    messagebox.showinfo("Выберите файл", "Папка с шаблонами НЕ найдена!!!")
    folder_docx = fd.askdirectory(title='Выберите папку с шаблонами')
    
today = datetime.now()
tomorrow = today + timedelta(days=1)
tomorrow = tomorrow.date().strftime("%d.%m.%Y")

#Проверка наличие файла разбор МО
file_xlsx =r'C:\Users\User\Емиас\Корректировка ЕГИСЗ\разбор МО от 31.10.22.xlsx'
if os.path.isfile(file_xlsx):
    print ('Файл Разбор в МО найден')
else:
    print ('Файл "Разбор в МО" НЕ найден!!!')
    print ('Выберите файл с базой для рассыллки ')
    messagebox.showinfo('Выберите файл', 'Файл "Разбор в МО" НЕ найден!!!')
    file_xlsx = fd.askopenfilename(title='Выберите Файл "Разбор в МО"', initialdir=r'C:\Users\User\Емиас\Корректировка ЕГИСЗ')


root = tk.Tk()
root.title("Для разбора в МО")
root.geometry('400x300')
root["bg"] = "#fff"


send_list = tk.Button(text="Создать лист для отправки в МО",
                    command=create_send_list, background="#fff", foreground="#3b3e41",
                    padx="30", pady="15", font="15")

sep_xlsx = tk.Button(text="Разделение на МО",
                    command=sep_base_mo, background="#fff", foreground="#3b3e41",
                    padx="30", pady="15", font="15")

join_xlsx = tk.Button(text="Внесение отработи МО",
                    command=join_xlsx_mo, background="#fff", foreground="#3b3e41",
                    padx="30", pady="15", font="15")


send_list.pack(padx="30", pady="15")
sep_xlsx.pack(padx="30", pady="15")
join_xlsx.pack(padx="30", pady="15")


root.mainloop()
